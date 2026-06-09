#!/usr/bin/env python3
"""
Monta US_CPI.xlsx com duas abas:
  - US_CPI         : indices mensais CPI-U (SA onde disponivel, NSA em azul)
  - US_CPI_Weights : Relative Importance anual por categoria

Fonte dos indices : BLS Public Data API v2 (series CUSR/CUUR).
Fonte dos pesos  : BLS Public Data API v2, parametro aspects=True nas
                   series NSA (CUUR*), aspecto "Relative Importance" de
                   dezembro de cada ano. Cobertura: dez/2012 - dez/2025.
                   Celulas anteriores a 2013 ficam em branco.

Uso:
    pip install requests openpyxl
    python build_us_cpi.py

Saida: US_CPI.xlsx na mesma pasta.
"""

import csv
import json
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------
API_KEY     = "cca642797061495f80be412007ec5263"
START_YEAR  = 2000
END_YEAR    = 2026
MAPPING_CSV = "cpi_mapping.csv"
OUTPUT      = "US_CPI.xlsx"

API_URL      = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
BATCH_SERIES = 50
BATCH_YEARS  = 20

MONTHS = {
    "M01": 1, "M02": 2, "M03": 3, "M04": 4, "M05": 5,  "M06": 6,
    "M07": 7, "M08": 8, "M09": 9, "M10": 10, "M11": 11, "M12": 12,
}
MON_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May",  6: "Jun",
            7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


# ============================================================
# MAPEAMENTO E API DE INDICES
# ============================================================

def load_mapping(path):
    rows = []
    with open(path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            r["indent"] = int(r["indent"]) if r["indent"] != "" else None
            rows.append(r)
    return rows


def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def fetch_block(series_ids, start, end):
    payload = {
        "seriesid": series_ids,
        "startyear": str(start),
        "endyear": str(end),
        "registrationkey": API_KEY,
    }
    resp = requests.post(API_URL, data=json.dumps(payload),
                         headers={"Content-type": "application/json"},
                         timeout=120)
    resp.raise_for_status()
    j = resp.json()
    if j.get("status") != "REQUEST_SUCCEEDED":
        raise RuntimeError(f"API status {j.get('status')}: {j.get('message')}")
    return j["Results"]["series"]


def fetch_all(series_ids):
    """Retorna dict: series_id -> {(year, month): float}."""
    data = {sid: {} for sid in series_ids}
    windows = []
    y = START_YEAR
    while y <= END_YEAR:
        y2 = min(y + BATCH_YEARS - 1, END_YEAR)
        windows.append((y, y2))
        y = y2 + 1

    total = len(list(chunk(series_ids, BATCH_SERIES))) * len(windows)
    done  = 0
    for batch in chunk(series_ids, BATCH_SERIES):
        for (ys, ye) in windows:
            done += 1
            print(f"  requisicao {done}/{total}: {len(batch)} series, {ys}-{ye}")
            series = fetch_block(batch, ys, ye)
            for s in series:
                sid = s["seriesID"]
                for it in s.get("data", []):
                    p = it["period"]
                    if p not in MONTHS:
                        continue
                    try:
                        val = float(it["value"])
                    except ValueError:
                        continue
                    data[sid][(int(it["year"]), MONTHS[p])] = val
            time.sleep(0.3)
    return data


# ============================================================
# PESOS (RELATIVE IMPORTANCE via API aspects)
# ============================================================

def _nsa_sid(sid):
    """Converte series_id SA (CUSR*) para NSA equivalente (CUUR*).
    Series ja NSA (CUUR*) sao retornadas sem alteracao."""
    return sid.replace("CUSR", "CUUR", 1) if sid.startswith("CUSR") else sid


def fetch_relative_importance(mapping):
    """
    Busca o aspecto 'Relative Importance' de dezembro de cada ano
    para todas as series NSA equivalentes do mapeamento.

    A BLS grava o RI no ponto de dados de dezembro de cada ano.
    Retorna dict: nsa_series_id -> {dec_year(int): ri_value(float)}.
    Cobertura tipica: dezembro 2012 ate dezembro 2025.
    """
    nsa_ids = list({_nsa_sid(m["series_id"]) for m in mapping if m["series_id"]})
    ri      = {sid: {} for sid in nsa_ids}

    windows = [(2000, 2019), (2020, END_YEAR)]
    total   = len(list(chunk(nsa_ids, BATCH_SERIES))) * len(windows)
    done    = 0

    for batch in chunk(nsa_ids, BATCH_SERIES):
        for (ys, ye) in windows:
            done += 1
            print(f"  RI req {done}/{total}: {len(batch)} series, {ys}-{ye}")
            payload = {
                "seriesid":       batch,
                "startyear":      str(ys),
                "endyear":        str(ye),
                "aspects":        True,
                "registrationkey": API_KEY,
            }
            resp = requests.post(API_URL, data=json.dumps(payload),
                                 headers={"Content-type": "application/json"},
                                 timeout=120)
            resp.raise_for_status()
            j = resp.json()
            if j.get("status") != "REQUEST_SUCCEEDED":
                print(f"    Aviso RI: {j.get('message')}")
                continue
            for s in j["Results"]["series"]:
                sid = s["seriesID"]
                for d in s.get("data", []):
                    if d["period"] != "M12":
                        continue
                    for asp in d.get("aspects", []):
                        if asp["name"] == "Relative Importance":
                            try:
                                ri[sid][int(d["year"])] = float(asp["value"])
                            except (ValueError, TypeError):
                                pass
            time.sleep(0.3)

    found = sum(1 for v in ri.values() if v)
    print(f"  Series NSA com RI: {found}/{len(nsa_ids)}")
    return ri


# ============================================================
# ESTILOS COMPARTILHADOS
# ============================================================

def _make_styles(font_name="Arial"):
    fn = font_name
    return {
        "title":      Font(name=fn, bold=True, size=12),
        "header":     Font(name=fn, bold=True, color="FFFFFF"),
        "header_fill":PatternFill("solid", fgColor="C00000"),
        "cat":        Font(name=fn),
        "sa":         Font(name=fn, color="000000"),
        "nsa":        Font(name=fn, color="0000FF"),
        "note":       Font(name=fn, italic=True, size=9, color="404040"),
        "fn":         fn,
    }


def _write_headers(ws, timeline, title_val, note_val, styles):
    fn   = styles["fn"]
    ws.cell(row=1, column=2, value=title_val).font = styles["title"]
    ws.cell(row=2, column=2, value=note_val).font  = styles["note"]

    ws.cell(row=4, column=1, value="Indent Level").font = styles["header"]
    ws.cell(row=4, column=1).fill = styles["header_fill"]
    ws.cell(row=4, column=3, value="Expenditure category").font = styles["header"]
    ws.cell(row=4, column=3).fill = styles["header_fill"]

    for i, (y, m) in enumerate(timeline):
        c = ws.cell(row=5, column=4 + i,
                    value=f"{MON_ABBR[m]}-{str(y)[2:]}")
        c.font      = styles["header"]
        c.fill      = styles["header_fill"]
        c.alignment = Alignment(horizontal="center")


def _set_dimensions(ws, timeline):
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 52
    for i in range(len(timeline)):
        ws.column_dimensions[
            ws.cell(row=5, column=4 + i).column_letter
        ].width = 9
    ws.freeze_panes = "D6"


# ============================================================
# CONSTRUCAO DAS ABAS
# ============================================================

def _write_index_sheet(ws, mapping, data, timeline, styles):
    _write_headers(ws, timeline,
        "Consumer Price Index for All Urban Consumers (CPI-U): "
        "U.S. city average, by expenditure category",
        "[1982-84=100, unless otherwise noted] | "
        "Seasonally adjusted indexes. Valores em azul = "
        "Not Seasonally Adjusted (SA indisponivel na fonte).",
        styles)

    r = 6
    for mp in mapping:
        ws.cell(row=r, column=1, value=mp["indent"]).font = styles["cat"]
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=mp["category"]).font = styles["cat"]
        sid   = mp["series_id"]
        basis = mp["basis"]
        if sid and sid in data:
            series   = data[sid]
            use_font = styles["nsa"] if basis == "NSA" else styles["sa"]
            for i, key in enumerate(timeline):
                if key in series:
                    cell = ws.cell(row=r, column=4 + i, value=series[key])
                    cell.font          = use_font
                    cell.number_format = "0.000"
        r += 1

    _set_dimensions(ws, timeline)


def _write_weights_sheet(ws, mapping, ri_by_nsa, timeline, styles):
    _write_headers(ws, timeline,
        "Consumer Price Index for All Urban Consumers (CPI-U): "
        "U.S. city average — Relative Importance by expenditure category",
        "Relative Importance (BLS API, aspecto de dezembro de cada ano, soma = 100). "
        "Para cada mes do ano AAAA o peso exibido e o de dezembro AAAA. "
        "Cobertura: dez/2012-dez/2025. Pre-2013: celulas em branco. "
        "Valores em azul = categoria NSA.",
        styles)

    r = 6
    for mp in mapping:
        ws.cell(row=r, column=1, value=mp["indent"]).font = styles["cat"]
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=mp["category"]).font = styles["cat"]
        sid   = mp["series_id"]
        basis = mp["basis"]
        if sid:
            nsa      = _nsa_sid(sid)
            use_font = styles["nsa"] if basis == "NSA" else styles["sa"]
            yr_map   = ri_by_nsa.get(nsa, {})
            if yr_map:
                for i, (y, m) in enumerate(timeline):
                    # usa o RI de dezembro do mesmo ano; se ausente, retrocede
                    for yr in range(y, y - 15, -1):
                        if yr in yr_map:
                            cell = ws.cell(row=r, column=4 + i, value=yr_map[yr])
                            cell.font          = use_font
                            cell.number_format = "0.000"
                            break
        r += 1

    _set_dimensions(ws, timeline)


# ============================================================
# WORKBOOK PRINCIPAL
# ============================================================

def build_workbook(mapping, data, ri_by_nsa):
    all_keys = set()
    for sid, series in data.items():
        all_keys.update(series.keys())
    if not all_keys:
        raise RuntimeError("Nenhum dado retornado pela API.")

    min_key = min(all_keys)
    max_key = max(all_keys)
    start   = (START_YEAR, 1)
    if min_key < start:
        min_key = start

    timeline = []
    y, m = min_key
    while (y, m) <= max_key:
        timeline.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    styles = _make_styles()
    wb     = Workbook()

    ws1 = wb.active
    ws1.title = "US_CPI"
    _write_index_sheet(ws1, mapping, data, timeline, styles)

    ws2 = wb.create_sheet(title="US_CPI_Weights")
    _write_weights_sheet(ws2, mapping, ri_by_nsa, timeline, styles)

    wb.save(OUTPUT)
    print(f"\nSalvo: {OUTPUT}")
    print(f"Periodo: {MON_ABBR[timeline[0][1]]}-{timeline[0][0]} "
          f"ate {MON_ABBR[timeline[-1][1]]}-{timeline[-1][0]} "
          f"({len(timeline)} meses)")
    cats = len([m for m in mapping if m["series_id"]])
    ri_years = sorted({yr for v in ri_by_nsa.values() for yr in v})
    print(f"Categorias: {cats}")
    if ri_years:
        print(f"Pesos (RI): dez/{ri_years[0]} - dez/{ri_years[-1]} "
              f"({len(ri_years)} anos)")


# ============================================================
# MAIN
# ============================================================

def main():
    mapping    = load_mapping(MAPPING_CSV)
    series_ids = [m["series_id"] for m in mapping if m["series_id"]]

    print(f"Buscando {len(series_ids)} series na API da BLS...")
    data = fetch_all(series_ids)
    got  = sum(1 for sid in series_ids if data[sid])
    print(f"Series com dados: {got}/{len(series_ids)}")

    print(f"\nBuscando pesos (Relative Importance) via API...")
    ri_by_nsa = fetch_relative_importance(mapping)

    build_workbook(mapping, data, ri_by_nsa)


if __name__ == "__main__":
    main()
